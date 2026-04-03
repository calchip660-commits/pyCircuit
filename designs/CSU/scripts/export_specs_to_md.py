#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export CSU binary specs (XLSX, DOCX, PDF) to Markdown under docs/converted/.

Run from repo root::

    python3 designs/CSU/scripts/export_specs_to_md.py

Requires: pandoc (for .docx), openpyxl, pypdf.
"""
from __future__ import annotations

import re
import subprocess
import sys
from datetime import date
from pathlib import Path

_REPO = Path(__file__).resolve().parents[3]
_DOCS = _REPO / "designs" / "CSU" / "docs"
_OUT = _DOCS / "converted"


def _md_escape_cell(x: object) -> str:
    if x is None:
        return ""
    s = str(x).replace("\n", " ").replace("|", "\\|")
    return s


def export_xlsx(path: Path) -> list[Path]:
    import openpyxl

    written: list[Path] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        safe = re.sub(r"[^\w\-]+", "_", sheet).strip("_") or "sheet"
        out = _OUT / f"SRC-01_xlsx_{safe}.md"
        lines: list[str] = [
            f"# SRC-01 — XLSX sheet `{sheet}`",
            "",
            f"**Source:** `{path.name}`  ",
            f"**Generated:** {date.today().isoformat()}  ",
            f"**Tool:** `designs/CSU/scripts/export_specs_to_md.py` (openpyxl)",
            "",
        ]
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            lines.append("*(empty sheet)*")
        else:
            ncols = max(len(r) for r in rows if r)
            # trim trailing all-None columns
            while ncols > 0 and all(
                (r[ncols - 1] if r and len(r) >= ncols else None) in (None, "")
                for r in rows
            ):
                ncols -= 1
            header = rows[0][:ncols] if rows else []
            lines.append("| " + " | ".join(_md_escape_cell(header[i]) for i in range(ncols)) + " |")
            lines.append("| " + " | ".join("---" for _ in range(ncols)) + " |")
            for r in rows[1:]:
                if r is None:
                    continue
                cells = [r[i] if i < len(r) else None for i in range(ncols)]
                if all(c is None or c == "" for c in cells):
                    continue
                lines.append("| " + " | ".join(_md_escape_cell(c) for c in cells) + " |")
        out.write_text("\n".join(lines) + "\n", encoding="utf-8")
        written.append(out)
    return written


def export_docx(path: Path) -> Path:
    media = _OUT / "SRC-07_media"
    media.mkdir(parents=True, exist_ok=True)
    out = _OUT / "SRC-07_linxcore950_csu_design_spec.md"
    cmd = [
        "pandoc",
        str(path),
        "-t",
        "markdown",
        "-o",
        str(out),
        f"--extract-media={media}",
    ]
    subprocess.run(cmd, check=True)
    header = (
        f"<!-- **Source:** `{path.name}` | **Generated:** {date.today().isoformat()} | pandoc --extract-media=SRC-07_media -->\n\n"
    )
    body = out.read_text(encoding="utf-8")
    out.write_text(header + body, encoding="utf-8")
    return out


def export_pdf_sample(path: Path, max_pages: int = 150) -> Path:
    from pypdf import PdfReader

    out = _OUT / "SRC-08_chi_architecture_spec_extract.md"
    reader = PdfReader(str(path))
    n = min(len(reader.pages), max_pages)
    chunks: list[str] = [
        "# SRC-08 — CHI architecture spec (partial text extract)",
        "",
        f"**Source:** `{path.name}` (Arm IHI0050H)  ",
        f"**Generated:** {date.today().isoformat()}  ",
        f"**Tool:** pypdf — **pages 1–{n} only** (full PDF is large; use original PDF for normative figures/tables).",
        "",
    ]
    for i in range(n):
        text = reader.pages[i].extract_text() or ""
        chunks.append(f"## PDF page {i + 1}\n\n```text\n{text.strip()}\n```\n")
    out.write_text("\n".join(chunks), encoding="utf-8")
    return out


def write_readme(xlsx_n: int, docx_ok: bool, pdf_ok: bool) -> None:
    _OUT.mkdir(parents=True, exist_ok=True)
    readme = _OUT / "README.md"
    readme.write_text(
        f"""# Converted CSU specifications (Markdown)

Binary sources under `designs/CSU/docs/` are exported here so agents and text tools can search and cite them.

| Artifact | Original | Converted output |
|----------|----------|------------------|
| SRC-01 | `CSU 接口Protocol_辅助设计输入.xlsx` | `SRC-01_xlsx_*.md` ({xlsx_n} sheets) |
| SRC-07 | `LinxCore950 CSU Design Specification-AI辅助设计输入.docx` | `SRC-07_linxcore950_csu_design_spec.md` + `SRC-07_media/` |
| SRC-08 | `IHI0050H_amba_chi_architecture_spec.pdf` | `SRC-08_chi_architecture_spec_extract.md` (first pages only) |

**Regenerate:** `python3 designs/CSU/scripts/export_specs_to_md.py` from repo root.

**DOCX export:** {"OK (pandoc)" if docx_ok else "skipped/failed"}  
**PDF extract:** {"OK (pypdf)" if pdf_ok else "skipped/failed"}

Figures in DOCX are referenced from `SRC-07_media/`; equation-heavy PDF pages may have imperfect text layout — always keep the original PDF for sign-off.
""",
        encoding="utf-8",
    )


def main() -> int:
    _OUT.mkdir(parents=True, exist_ok=True)
    xlsx_path = _DOCS / "CSU 接口Protocol_辅助设计输入.xlsx"
    docx_path = _DOCS / "LinxCore950 CSU Design Specification-AI辅助设计输入.docx"
    pdf_path = _DOCS / "IHI0050H_amba_chi_architecture_spec.pdf"

    if not xlsx_path.is_file():
        print("missing XLSX:", xlsx_path, file=sys.stderr)
        return 1

    written = export_xlsx(xlsx_path)
    print("xlsx ->", len(written), "markdown files")

    docx_ok = False
    if docx_path.is_file():
        try:
            export_docx(docx_path)
            docx_ok = True
            print("docx -> SRC-07_linxcore950_csu_design_spec.md")
        except Exception as e:  # noqa: BLE001
            print("docx export failed:", e, file=sys.stderr)

    pdf_ok = False
    if pdf_path.is_file():
        try:
            export_pdf_sample(pdf_path)
            pdf_ok = True
            print("pdf -> SRC-08_chi_architecture_spec_extract.md (partial)")
        except Exception as e:  # noqa: BLE001
            print("pdf export failed:", e, file=sys.stderr)

    write_readme(len(written), docx_ok, pdf_ok)
    print("wrote", _OUT / "README.md")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
